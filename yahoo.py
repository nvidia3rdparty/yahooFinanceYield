'''
Created on 6.2.2017
Version 0.1

License
You are free to use this to commercial of non-commercial use. 
Just credit me, Riku Pasonen as original source and take note pyhton license rules

Must be used with portfolio.xlsx file in the same folder
You can change the sheet names in portfolio but you must do same changes here also.
I have my stocks divided into three groups(REIT,Finance,General)
You must change the stock list in excel to match yours
Wont work with ETFs or mutual funds. Use tickers what yahoo finance uses.

'''
from yahoo_finance import Share

#this sets how many stocks are per sheet
max_rows_to_try=100

from openpyxl import load_workbook
wb2 = load_workbook('portfolio.xlsx')
print (wb2.get_sheet_names())
f1 = wb2['Finance']

colA = f1['A']
div=[]
row_range = f1[3:max_rows_to_try]
for row in f1.iter_rows(min_row=3, max_col=1, max_row=max_rows_to_try):
    for cell in row: 
        if cell.value!=None:
            try:
                dividend=Share(cell.value).get_dividend_yield()
                div.append([dividend,str(cell.row)])
            
            except Exception: 
                pass
           
      
        else:
            
            break
        
        print(cell.value+" "+str(cell.row))
print(div[1])
for unit in div:
    if unit[0]!=None:
        f1['B'+unit[1]] = unit[0]
    else:
        continue

f2 = wb2['REIT']
colA = f2['A']
div=[]
row_range = f2[3:max_rows_to_try]
for row in f2.iter_rows(min_row=3, max_col=1, max_row=max_rows_to_try):
    for cell in row: 
        if cell.value!=None:
            try:
                dividend=Share(cell.value).get_dividend_yield()
                div.append([dividend,str(cell.row)])
            
            except Exception: 
                pass
           
      
        else:
            
            break
        print(cell.value+" "+str(cell.row))

for unit in div:
    if unit[0]!=None:
        f2['B'+unit[1]] = unit[0]
    else:
        continue

f3 = wb2['General']
colA = f3['A']
div=[]
row_range = f3[3:max_rows_to_try]
for row in f3.iter_rows(min_row=3, max_col=1, max_row=max_rows_to_try):
    for cell in row: 
        if cell.value!=None:
            try:
                dividend=Share(cell.value).get_dividend_yield()
                div.append([dividend,str(cell.row)])
            
            except Exception: 
                pass
           
      
        else:
            
            break
        print(cell.value+" "+str(cell.row))

for unit in div:
    if unit[0]!=None:
        f3['B'+unit[1]] = unit[0]
    else:
        continue


    
wb2.save(filename = 'portfolio.xlsx')
print("All done!")


    
